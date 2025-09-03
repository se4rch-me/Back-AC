import json
import os
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import PatternFill, Border, Side, Font

from .image_generator import create_connections_table_image

# --- LA "RECETA" MAESTRA DEL REPORTE ---
CELL_MAPPING = {
    'fecha': {'type': 'direct', 'cell': 'D3'},
    'consecutivo': {'type': 'direct', 'cell': 'H3'},
    'direccion': {'type': 'direct', 'cell': 'D4'},
    'levanto': {'type': 'direct', 'cell': 'D5'},
    'pozo_numero': {'type': 'direct', 'cell': 'D6'},
    'cilindro_cual': {'type': 'direct', 'cell': 'H52'},
    'observaciones': {'type': 'direct', 'cell': 'M80'},
    'tipo_sistema': {'type': 'options', 'values': {'Aguas Lluvia': 'D9', 'Aguas Residuales': 'F9', 'Combinado': 'H9'}},
    'tipo_pozo': {'type': 'options', 'values': {'Pozo': 'D11', 'Camara': 'F11', 'Alivio': 'H11'}},
    'tapa_existe': {'type': 'options', 'values': {'Si': 'D16', 'No': 'F16'}},
    'tapa_tipo': {'type': 'options', 'values': {'Ferroconcreto': 'D18', 'Concreto': 'F18', 'Hierro sin Bisagra': 'H18', 'Hierro con bisagra': 'D19', 'Tapa Seguridad': 'F19', 'Tapa en fibra': 'H19'}},
    'tapa_estado': {'type': 'options', 'values': {'Bueno': 'D21', 'Regular': 'F21', 'Malo': 'H21'}},
    'tapa_diagnostico': {'type': 'options', 'values': {'Cambiar': 'D23', 'Reparar': 'F23', 'No Requiere': 'H23'}},
    'cargue_existe': {'type': 'options', 'values': {'Si': 'D28', 'No': 'F28'}},
    'cargue_estado': {'type': 'options', 'values': {'Bueno': 'D30', 'Regular': 'F30', 'Malo': 'H30', 'Grietas': 'D31', 'Partido': 'F31', 'Hundido': 'H31'}},
    'cargue_diagnostico': {'type': 'options', 'values': {'Cambiar': 'D33', 'Reparar': 'F33', 'No Requiere': 'H33'}},
    'cono_existe': {'type': 'options', 'values': {'Si': 'D38', 'No': 'F38'}},
    'cono_estado': {'type': 'options', 'values': {'Bueno': 'D40', 'Regular': 'F40', 'Malo': 'H40', 'Grietas': 'D41', 'Partido': 'F41', 'Hundido': 'H41'}},
    'cono_diagnostico': {'type': 'options', 'values': {'Cambiar': 'D43', 'Reparar': 'F43', 'No Requiere': 'H43'}},
    'cilindro_material': {'type': 'options', 'values': {'Mamposteria': 'D48', 'Concreto': 'F48', 'GRP': 'H48'}},
    'cilindro_estado': {'type': 'options', 'values': {'Bueno': 'D50', 'Regular': 'F50', 'Malo': 'H50', 'Grietas': 'D51', 'Partido': 'F51', 'Huecos': 'H51', 'Sin Pañete': 'D52', 'Otro': 'F52'}},
    'cilindro_diagnostico': {'type': 'options', 'values': {'Cambiar': 'D54', 'Reparar': 'F54', 'No Requiere': 'H54'}},
    'canuela_estado': {'type': 'options', 'values': {'Bueno': 'D59', 'Regular': 'F59', 'Malo': 'H59', 'Sedimentada': 'D60', 'Desgastada': 'F60', 'Socavacion': 'H60'}},
    'canuela_diagnostico': {'type': 'options', 'values': {'Cambiar': 'D62', 'Reparar': 'F62', 'No Requiere': 'H62'}},
    'escalones_existe': {'type': 'options', 'values': {'Si': 'D67', 'No': 'F67'}},
    'escalones_tipo': {'type': 'options', 'values': {'Escalones': 'D69', 'Ladrillos': 'F69'}},
    'escalones_estado': {'type': 'options', 'values': {'Bueno': 'D71', 'Regular': 'F71', 'Malo': 'H71', 'Doblados': 'D72', 'Faltan': 'F72', 'Corroidos': 'H72'}},
    'escalones_diagnostico': {'type': 'options', 'values': {'Cambiar': 'D74', 'Reparar': 'F74', 'No Requiere': 'H74'}},
    'estado_general_pozo': {'type': 'options', 'values': {'Infiltracion': 'D79', 'Represado': 'F79', 'Con basura': 'H79', 'Raices': 'D80', 'Fuera de Servicio': 'F80', 'Lleno de tierra': 'H80'}}
}

def fill_sheet(sheet, record):
    """Rellena una hoja de cálculo con los datos de un registro."""
    print(f"  - Rellenando hoja para el pozo '{record.get('pozo_numero')}'...")

    # 1. Relleno de celdas
    for field, mapping in CELL_MAPPING.items():
        value = record.get(field, '')

        if field == 'observaciones':
            cell_coord = mapping['cell']
            text = " " + str(value)
            
            if len(text) > 44:
                processed_text = text[:44]
            else:
                processed_text = text.ljust(44, '_')

            cell = sheet[cell_coord]
            cell.value = processed_text
            cell.font = cell.font.copy(underline='single')

        elif value is not None and value != '':
            if mapping['type'] == 'direct':
                sheet[mapping['cell']] = value
            elif mapping['type'] == 'options':
                cell_to_mark = mapping['values'].get(str(value))
                if cell_to_mark:
                    sheet[cell_to_mark] = 'X'

    # 2. Lógica de conexiones para crear imagen
    try:
        connections_json = record.get('conexiones', '[]')
        connections = json.loads(connections_json)
        if connections:
            sheet['D7'] = connections[0].get('cota_razante')
            
            width_px = 338

            table_img_buffer = create_connections_table_image(connections, target_width_px=width_px)
            img_tabla = OpenpyxlImage(table_img_buffer)
            sheet.add_image(img_tabla, 'M56')
    except Exception as e:
        print(f"  - Advertencia al procesar conexiones: {e}")

    # 3. Lógica para la imagen del esquema
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(script_dir)
        placeholder_path = os.path.join(project_root, 'esquema_placeholder.png')

        if os.path.exists(placeholder_path):
            img_placeholder = OpenpyxlImage(placeholder_path)
            sheet.add_image(img_placeholder, 'M2')
            print("  - Insertando placeholder para el esquema del pozo.")
        else:
            print(f"  - ¡ERROR CRÍTICO! No se encontró el archivo placeholder en la ruta esperada: '{placeholder_path}'.")

    except Exception as e:
        print(f"  - Error al procesar la imagen del esquema: {e}")

    # 4. Aplicar estilos de marco y fondo
    try:
        gray_fill = PatternFill(start_color="ADADAD", end_color="ADADAD", fill_type="solid")
        for row in sheet.iter_rows():
            for cell in row:
                if cell.row > 82 or cell.column > 14:
                    cell.fill = gray_fill
        print("  - Aplicando color de fondo por defecto.")

        blue_medium_side = Side(border_style="medium", color="0000FF")
        
        for row_idx in range(1, 83):
            cell = sheet.cell(row=row_idx, column=14)
            cell.border = cell.border.copy(right=blue_medium_side)

        for col_idx in range(1, 15):
            cell = sheet.cell(row=82, column=col_idx)
            cell.border = cell.border.copy(bottom=blue_medium_side)
        print("  - Aplicando bordes al marco.")

    except Exception as e:
        print(f"  - Error al aplicar estilos de marco y fondo: {e}")

    print(f"  - Hoja para el pozo '{record.get('pozo_numero')}' rellenada.")
