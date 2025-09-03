import openpyxl
import io
import os
import time
from copy import deepcopy

from google_clients import get_credentials
from report_generation.sheets_handler import get_pending_records, update_record_status
from report_generation.drive_handler import download_master_report, update_master_report
from report_generation.excel_handler import fill_sheet

def main():
    """Función principal que orquesta la generación de reportes."""
    print("--- Iniciando Proceso de Generación de Reportes (Modular) ---")
    master_report_path = None  # Inicializar la variable de ruta
    try:
        # 1. Verificar credenciales
        get_credentials()
        print("Autenticación con Google verificada.")

        # 2. Obtener registros pendientes
        worksheet, pending_records, _ = get_pending_records()
        if not pending_records:
            print("No hay registros pendientes. Finalizando.")
            return

        # 3. Lógica de espera y descarga del REPORTE MAESTRO
        print("Esperando 7 segundos para mitigar posible caché de Google Drive...")
        time.sleep(7)
        master_report_path = download_master_report()
        workbook = openpyxl.load_workbook(master_report_path)
        
        # 4. Usar la primera hoja del libro de trabajo descargado como plantilla
        template_sheet = workbook[workbook.sheetnames[0]]
        print(f"Usando la hoja '{template_sheet.title}' como plantilla.")

        # 5. Procesar cada registro pendiente
        processed_rows = []
        for row_number, record in pending_records:
            base_name = str(record.get('pozo_numero', f"Fila_{row_number}"))
            
            # Lógica para nombres de hoja únicos
            sheet_title = base_name
            counter = 2
            while sheet_title in workbook.sheetnames:
                sheet_title = f"{base_name}({counter})"
                counter += 1

            print(f"\nProcesando Pozo: {base_name} -> Creando hoja: '{sheet_title}'")
            
            # Copiar la hoja de plantilla DENTRO del mismo libro de trabajo
            new_sheet = workbook.copy_worksheet(template_sheet)
            new_sheet.title = sheet_title
            new_sheet.sheet_view.showGridLines = False  # Ocultar líneas de cuadrícula

            # Copiar las imágenes una por una con deepcopy
            if hasattr(template_sheet, '_images') and template_sheet._images:
                for img in template_sheet._images:
                    new_img = deepcopy(img)
                    new_sheet.add_image(new_img, new_img.anchor)

            fill_sheet(new_sheet, record)
            
            processed_rows.append(row_number)

        # 6. Guardar y Subir
        output_buffer = io.BytesIO()
        workbook.save(output_buffer)
        output_buffer.seek(0)
        update_master_report(output_buffer)

        # 7. Actualizar estado en Sheets
        update_record_status(worksheet, processed_rows)

    except Exception as e:
        print(f"--- ¡Ocurrió un error inesperado! ---")
        print(f"Error: {e}")
    
    finally:
        # 8. Limpiar el archivo temporal
        if master_report_path and os.path.exists(master_report_path):
            os.remove(master_report_path)
            print(f"Archivo temporal eliminado: {master_report_path}")

    print("\n--- Proceso Finalizado ---")

if __name__ == '__main__':
    main()
